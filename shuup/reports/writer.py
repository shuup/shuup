# -*- coding: utf-8 -*-
# This file is part of Shuup.
#
# Copyright (c) 2012-2021, Shuup Commerce Inc. All rights reserved.
#
# This source code is licensed under the OSL-3.0 license found in the
# LICENSE file in the root directory of this source tree.
import bleach
import csv
import six
from babel.dates import format_datetime
from datetime import datetime
from decimal import Decimal
from django.core.serializers.json import DjangoJSONEncoder
from django.db import models
from django.http import HttpResponse
from django.template.defaultfilters import floatformat
from django.template.loader import render_to_string
from django.utils.encoding import smart_text
from django.utils.functional import Promise
from django.utils.html import conditional_escape, escape
from django.utils.safestring import mark_safe
from django.utils.timezone import now
from io import StringIO
from pprint import pformat
from six import BytesIO

from shuup.apps.provides import get_provide_objects
from shuup.utils.django_compat import force_text
from shuup.utils.i18n import format_money, get_current_babel_locale, get_locally_formatted_datetime
from shuup.utils.money import Money
from shuup.utils.pdf import render_html_to_pdf

try:
    import openpyxl
except ImportError:
    openpyxl = None


REPORT_WRITERS_MAP = {}


class ReportWriter(object):
    content_type = None
    extension = None
    inline = False  # Implementation-dependent
    writer_type = "base"

    def __init__(self):
        self.title = ""

    def __unicode__(self):
        return self.writer_type

    def __str__(self):
        return self.writer_type

    def write_heading(self, text):
        raise NotImplementedError("Error! Not implemented: `ReportWriter` -> `write_heading()`.")

    def write_text(self, text):
        raise NotImplementedError("Error! Not implemented: `ReportWriter` -> `write_text()`.")

    def write_data_table(self, report, report_data, has_totals=True):
        raise NotImplementedError("Error! Not implemented: `ReportWriter` -> `write_data_table()`.")

    def write_template(self, template_name, env):
        raise NotImplementedError("Error! Not implemented: `ReportWriter` -> `write_template()`.")

    def next_page(self):
        pass

    def get_rendered_output(self):
        raise NotImplementedError("Error! Not implemented: `ReportWriter` -> `get_rendered_output()`.")

    def _render_report(self, report):
        if not report.rendered:
            report_data = report.get_data()
            self.write_heading(
                "{title} {start} - {end}".format(
                    title=report.title,
                    start=format_datetime(report_data["start"], format="short", locale=get_current_babel_locale()),
                    end=format_datetime(report_data["end"], format="short", locale=get_current_babel_locale()),
                )
            )
            report.ensure_texts()
            self.write_data_table(report, report_data["data"], has_totals=report_data["has_totals"])

        return self.get_rendered_output()

    def render_report(self, report, inline=False):
        """
        Renders given report

        :param report:
        :type report:
        :return:
        :rtype:
        """
        self.inline = inline
        rendered_output = self._render_report(report)
        if self.writer_type == "html":
            output = mark_safe(rendered_output)
        else:
            output = mark_safe("<pre>%s</pre>" % escape(rendered_output))
        return output

    def get_response(self, report):
        """
        Returns downloadable file response

        :param report:
        :type report:
        :return:
        :rtype:
        """
        response = HttpResponse(self._render_report(report), content_type=self.content_type)
        if report.filename_template:
            response["Content-Disposition"] = "attachment; filename=%s" % self.get_filename(report)
        return response

    def get_filename(self, report):
        fmt_data = dict(report.options, time=now().isoformat())
        return "%s%s" % ((report.filename_template % fmt_data).replace(":", "_"), self.extension)


def format_data(data, format_iso_dates=False, format_money_values=False):
    if data is None:
        return ""

    elif isinstance(data, Money):
        if format_money_values:
            return format_money(data)
        return float(data.as_rounded().value)

    elif isinstance(data, Decimal):
        exponent = abs(data.normalize().as_tuple().exponent)
        # Limit the amount of decimals to 10
        return floatformat(data, min(exponent, 10))

    elif callable(data):
        return force_text(data())

    elif isinstance(data, Promise):
        return force_text(data)

    elif isinstance(data, models.Model):
        return force_text(data)

    elif isinstance(data, datetime):
        if format_iso_dates:
            return data.isoformat()
        return get_locally_formatted_datetime(data)

    return data


def remove_unsafe_chars(data):
    if isinstance(data, str):
        return "".join([char for char in data if char not in ("=", "+", "-")])

    return data


class CSVReportWriter(ReportWriter):
    content_type = "text/csv"
    extension = ".csv"
    writer_type = "csv"

    def __init__(self):
        super(CSVReportWriter, self).__init__()
        self.data = []

    def write_heading(self, text):
        self.data.append([text])

    def write_data_table(self, report, report_data, has_totals=True):
        self.data.append([c["title"] for c in report.schema])
        for datum in report_data:
            datum = report.read_datum(datum)
            self.data.append([format_data(remove_unsafe_chars(data), format_iso_dates=True) for data in datum])

        if has_totals:
            for datum in report.get_totals(report_data):
                datum = report.read_datum(datum)
                self.data.append([format_data(remove_unsafe_chars(data)) for data in datum])

    def get_rendered_output(self):
        f = StringIO()
        csv.writer(f).writerows(self.data)
        return f.getvalue().encode()


class ExcelReportWriter(ReportWriter):
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    extension = ".xlsx"
    writer_type = "excel"

    def __init__(self):
        super(ExcelReportWriter, self).__init__()
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active

    def next_page(self):
        self.worksheet = self.workbook.create_sheet()

    def _convert_row_to_string(self):
        # change the format of the last row to string if that is a formula
        for cell in self.worksheet[self.worksheet.max_row]:
            if cell.data_type == "f":
                cell.data_type = "s"

    def write_data_table(self, report, report_data, has_totals=True):
        self.worksheet.append([c["title"] for c in report.schema])
        for datum in report_data:
            datum = report.read_datum(datum)
            self.worksheet.append([format_data(remove_unsafe_chars(data)) for data in datum])
            self._convert_row_to_string()

        if has_totals:
            for datum in report.get_totals(report_data):
                datum = report.read_datum(datum)
                self.worksheet.append([format_data(remove_unsafe_chars(data)) for data in datum])
                self._convert_row_to_string()

    def write_page_heading(self, text):
        self.worksheet.append([text])
        self.worksheet.append([""])

    def write_heading(self, text):
        self.worksheet.append([text])

    def write_text(self, text):
        self.worksheet.append(text)

    def get_rendered_output(self):
        bio = BytesIO()
        self.workbook.save(bio)
        return bio.getvalue()


class HTMLReportWriter(ReportWriter):
    content_type = "text/html; charset=UTF-8"
    extension = ".html"
    writer_type = "html"

    INLINE_TEMPLATE = """
<body>
    %(body)s
</body>
    """.strip()

    TEMPLATE = (
        """
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>%(title)s</title>
%(extrahead)s
<style type="text/css">%(style)s</style>
</head>""".strip()
        + INLINE_TEMPLATE
        + """</html>"""
    )

    # styles = """@page { prince-shrink-to-fit: auto }""".strip()
    styles = """
@page { size: 8.5in 11in }
body { max-width: 95%}
table {
  table-layout: fixed;
  width: 100%;
}
table, th, td {
  word-break: break-all;
  word-wrap: break-word;
}
    """.strip()

    extra_header = ""

    def __init__(self):
        super(HTMLReportWriter, self).__init__()
        self.output = []

    def _w_raw(self, content):
        self.output.append(mark_safe(content))

    def _w(self, content):
        self.output.append(bleach.clean(str(format_data(content, format_money_values=True)), strip=True))

    def _w_tag(self, tag, content):
        self._w_raw("<%s>" % tag)
        self._w(content)
        self._w_raw("</%s>" % tag)

    def next_page(self):
        self._w_raw("<hr>")

    def write_data_table(self, report, report_data, has_totals=True):
        self._w_raw('<table class="table table-striped table-bordered">')
        self._w_raw("<thead><tr>")
        for c in report.schema:
            self._w_tag("th", c["title"])
        self._w_raw("</tr></thead>")
        self._w_raw("<tbody>")

        for datum in report_data:
            datum = report.read_datum(datum)
            self._w_raw("<tr>")
            for d in datum:
                self._w_tag("td", d)
            self._w_raw("</tr>")

        if has_totals:
            self._w_raw("<tr>")
            for d in report.read_datum(report.get_totals(report_data)):
                self._w_tag("td", d)
            self._w_raw("</tr>")

        self._w_raw("</tbody></table>")

    def write_page_heading(self, text):
        self._w_tag("h1", text)

    def write_heading(self, text):
        self._w_tag("h2", text)

    def write_text(self, text):
        self._w(text)

    def write_tag(self, tag, text):
        self._w_tag(tag, text)

    def get_rendered_output(self):
        body = "".join(conditional_escape(smart_text(piece)) for piece in self.output)
        styles = self.styles
        extrahead = self.extra_header or ""

        if self.inline:
            template = self.INLINE_TEMPLATE
        else:
            template = self.TEMPLATE

        html = template % {"title": self.title, "body": body, "style": styles, "extrahead": extrahead}
        if not self.inline:
            html = html.encode("UTF-8")
        return html

    def set_extra(self, extrastring):
        self.extra_header = extrastring

    def set_style(self, stylestring):
        self.styles = stylestring

    def write_template(self, template_name, env):
        data = render_to_string(template_name, env)
        if isinstance(data, str):
            data = data.decode("UTF-8")
        self._w_raw(data)


class PDFReportWriter(HTMLReportWriter):
    content_type = "application/pdf"
    extension = ".pdf"
    writer_type = "pdf"

    def get_rendered_output(self):
        html = HTMLReportWriter.get_rendered_output(self)
        pdf_response = render_html_to_pdf(html)
        return pdf_response.content


class JSONReportWriter(ReportWriter):
    content_type = "application/json"
    extension = ".json"
    writer_type = "json"

    def __init__(self):
        super(JSONReportWriter, self).__init__()
        self.data = {}

    def write_data_table(self, report, report_data, has_totals=True):
        table = {
            "columns": report.schema,
            "data": [
                dict(
                    (c["key"], format_data(val, format_iso_dates=True))
                    for (c, val) in zip(report.schema, report.read_datum(datum))
                )
                for datum in report_data
            ],
        }

        if has_totals:
            table["totals"] = report.get_totals(report_data)

        self.data.setdefault("tables", []).append(table)

    def write_heading(self, text):
        self.data["heading"] = text

    def write_text(self, text):
        pass

    def get_rendered_output(self):
        return DjangoJSONEncoder(indent=4).encode(self.data)


class PprintReportWriter(JSONReportWriter):
    content_type = "text/plain"
    extension = ".txt"
    writer_type = "pprint"

    def get_rendered_output(self):
        return pformat(self.data)


class ReportWriterPopulator(object):
    """
    A class which populates the report writers map.
    """

    report_writers_map = {}

    def populate(self):
        """
        Iterate over all report_writer_populator provides to fill/update the report writer map
        """
        for report_writer_populator_func in get_provide_objects("report_writer_populator"):
            report_writer_populator_func(self)

    def register(self, writer_name, writer_class):
        """
        Register a report writer for use.

        If a writer with same name already exists, it will be overwriten.

        :type writer_name: str
        :param writer_name: the unique name of the writer
        :type writer_class: ReportWriter
        :param writer_class: the report writer class
        """
        self.report_writers_map[writer_name] = writer_class

    @property
    def populated_map(self):
        """ Returns the populated map. """
        return self.report_writers_map


def get_writer_names():
    """ Get the registered writer names. """
    return set([k for k, v in six.iteritems(REPORT_WRITERS_MAP) if v])


def get_writer_instance(writer_name):
    """
    Get a report writer instance by name.

    :type writer_name: str
    :param writer_name: the name of the report writer
    :rtype: ReportWriter
    """
    writer = REPORT_WRITERS_MAP[writer_name]()
    assert isinstance(writer, ReportWriter)
    return writer


def populate_default_writers(writer_populator):
    """
    Populate the default report writers.

    :type writer_populator: ReportWriterPopulator
    """
    writer_populator.register("html", HTMLReportWriter)
    writer_populator.register("pdf", PDFReportWriter)
    writer_populator.register("json", JSONReportWriter)
    writer_populator.register("pprint", PprintReportWriter)
    writer_populator.register("csv", CSVReportWriter)

    if openpyxl:
        writer_populator.register("excel", ExcelReportWriter)


# populate them
populator = ReportWriterPopulator()
populator.populate()
REPORT_WRITERS_MAP.update(populator.populated_map)
